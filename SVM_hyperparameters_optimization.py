import scipy.io as sio
import os.path
import random
import numpy as np
import pandas as pd
from sklearn.model_selection import train_test_split, RandomizedSearchCV, GridSearchCV
from sklearn import svm
from sklearn.metrics import classification_report, confusion_matrix, accuracy_score
import seaborn as sns
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook


# take number random files since there are more CR than Harlem
def data_adq(group, number):

    path = os.listdir('./data/' + group + '/')
    files = list()
    for i in range(number):
        file = random.choice(path)
        while files.count(file):
            file = random.choice(path)

        files.append(file)
    return group, files


# combine feature vectors and add group classification
def data_prep(group, files, mode):
    # change work directory for simplicity
    os.chdir('./data/' + group + '/')
    feat_vect = list()
    # loop through files getting feature vector and adding the corresponding group
    for file in files:
        mat = sio.loadmat(file)
        theta = mat.get('theta')
        class_group = np.concatenate(([int(group[11])], theta.flatten()))
        feat_vect.append(class_group)
    data_save(group, mode, feat_vect)
    # return to home directory
    os.chdir('../../..')


# save the data in a file so the same dataset could be used for further tests
def data_save(group, mode, feat_vect):
    save_to = '../' + str(group[11]) + str(mode) + '.csv'
    np.savetxt(save_to, feat_vect, delimiter=",")


# manipulate the data to remove the labels
def features_manipulation(file1, file2, file3, file4):
    features1 = pd.read_csv(file1, header=None, delimiter=",")
    features2 = pd.read_csv(file2, header=None, delimiter=",")
    features = pd.concat([features1, features2])
    features.rename(columns={0: 'class'}, inplace=True)

    # Labels are the values we want to predict
    labels = np.array(features['class'])

    # Remove the labels from the features
    features = features.drop('class', axis=1)

    # Saving feature names for later use
    feature_list = list(features.columns)
    # Convert to numpy array
    features = np.array(features)

    features3 = pd.read_csv(file3, header=None, delimiter=",")
    features4 = pd.read_csv(file4, header=None, delimiter=",")
    features_test = pd.concat([features3, features4])
    features_test.rename(columns={0: 'class'}, inplace=True)

    # Labels are the values we want to predict
    labels_test = np.array(features_test['class'])

    # Remove the labels from the features
    features_test = features_test.drop('class', axis=1)

    # Convert to numpy array
    features_test = np.array(features_test)

    return features, labels, feature_list, features_test, labels_test


# split the data into train and test
# since there is previous train and test predefined it is forced to simplify
def split_data(features_train, features_test, labels_train, labels_test):
    # Split the data into training and testing sets
    train_features, _, train_labels, _ = train_test_split(features_train, labels_train, test_size=0.0001,
                                                          random_state=32)
    _, test_features, _, test_labels = train_test_split(features_test, labels_test, test_size=0.98,
                                                        random_state=32)
    # print('Training Features Shape:', train_features.shape)
    # print('Training Labels Shape:', train_labels.shape)
    # print('Testing Features Shape:', test_features.shape)
    # print('Testing Labels Shape:', test_labels.shape)
    return train_features, test_features, train_labels, test_labels


def create_SVM(model, train_features, test_features, train_labels, test_labels):
    # Train the model on training data
    model.fit(train_features, train_labels)
    # Use the forest's predict method on the test data
    predictions = model.best_estimator_.predict(test_features)

    #    sns.heatmap(table)
    #    plt.show()
    best_params = model.best_params_
    best_params['acc'] = accuracy_score(test_labels, predictions)

    return best_params


def random_search():

    svm_model = svm.SVC()
    # param_rs = {'kernel': ['linear', 'poly', 'rbf', 'sigmoid', 'precomputed'],
    #                  'C': list(np.linspace(2, 10, num=15, dtype = int)),
    #                  'degree': list(np.linspace(3, 100, num=15, dtype = int)),
    #                  'gamma': list(np.linspace(5, 500, num=13, dtype = float)) + ['scale', 'auto']
    #
    #             }
    param_rs = {'kernel': ['rbf'],
                #'C': list(np.linspace(2, 10, num=15, dtype=int)),
                'degree': list(np.linspace(3, 20, num=15, dtype=int)),
                'gamma': list(np.linspace(300, 600, num=30, dtype=float)) + ['scale', 'auto']

                }
    model = RandomizedSearchCV(estimator=svm_model, param_distributions=param_rs, n_iter=11, random_state=32)


    return model


def grid_search(best_params):
    print(best_params['gamma'])
    if best_params['degree'] < 11:
        best_params['degree'] = 11
    #
    # if best_params['C'] < 3:
    #     best_params['C'] = 3

    if type(best_params['gamma']) is float:
        gamma_array = [best_params['gamma']-2,best_params['gamma'], best_params['gamma']+2]
    else:
        gamma_array = [best_params['gamma']]


    param_gs = {'kernel': [best_params['kernel']],
                      #'C': [best_params['C'] - 2,
                      #                 best_params['C'],
                      #                 best_params['C'] + 2],
                      'degree': [best_params['degree'] - 10,
                                       best_params['degree'],
                                       best_params['degree'] + 10],

                     'gamma': gamma_array
                     }

    svm_model = svm.SVC()
    model = GridSearchCV(estimator=svm_model, param_grid=param_gs)

    return model



def array_to_Excel_col(dict, row, col, ws):
    i = 0
    for value in dict.values():
        e = ws.cell(row=row, column=col+i)
        if value is None:
            e.value = 'None'
        else:
            e.value = value
        i += 1


def params_to_Excel(dic_params,  row, col, ws):
    i = 0
    for params in dic_params:
        #print(params)
        array_to_Excel_col(params, row+i, col, ws)
        i += 1


if __name__ == '__main__':
    n_test = [41,51,40,44,44]
    n_train = [126,121,125,121,114]

    partitions = ['partition1/','partition2/','partition3/','partition4/', 'partition5/']

    wb = load_workbook('acc_results2.xlsx')
    ws1 = wb['SVM Random search']
    ws2 = wb['SVM Grid search']
    m = 0
    for k in range(0, len(partitions)):
        print('k =', k)
        j = 0
        configs_rs = []
        configs_gs = []

        for i in range(0, 6):
            print('i =', i)

            group_class_train, mat_files = data_adq(partitions[k]+'1-cr_train', n_train[k])
            data_prep(group_class_train, mat_files, 'train')
            group_class_test, mat_files = data_adq(partitions[k]+'1-cr_test', n_test[k])
            data_prep(group_class_test, mat_files, 'test')

            group_class2, mat_files2 = data_adq(partitions[k]+'2-hr_train', n_train[k])
            data_prep(group_class2, mat_files2, 'train')
            group_class2_test, mat_files2 = data_adq(partitions[k]+'2-hr_test', n_test[k])
            data_prep(group_class2_test, mat_files2, 'test')

            feature_noclass, label_class, feature_list, features_test, labels_test = features_manipulation(
                './data/' + partitions[k] + '1train.csv', './data/' + partitions[k] + '2train.csv', './data/' + partitions[k] +'1test.csv', './data/' + partitions[k] +'2test.csv')
            train_features, test_features, train_labels, test_labels = split_data(feature_noclass, features_test,
                                                                                  label_class, labels_test)

            # Grid Search based on Random Search
            model_rs = random_search()
            conf_bestparams_rs = create_SVM(model_rs, train_features, test_features, train_labels, test_labels)
            print('random search done')
            model_gs = grid_search(conf_bestparams_rs)
            conf_bestparams_gs = create_SVM(model_gs, train_features, test_features, train_labels, test_labels)
            print('grid search done')
            configs_rs.append(conf_bestparams_rs)
            configs_gs.append(conf_bestparams_gs)

            j += 1

        row = 3 + m
        col = 2
        print('RS ', configs_rs)
        print('GS ', configs_gs)
        params_to_Excel(configs_rs, row, col, ws1)
        params_to_Excel(configs_gs, row, col, ws2)
        m +=10

    wb.save('acc_results.xlsx')




