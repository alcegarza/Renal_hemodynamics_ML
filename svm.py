import scipy.io as sio
import os.path
import random
import numpy as np
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn import svm
from sklearn.metrics import classification_report, confusion_matrix, accuracy_score
from openpyxl import load_workbook, Workbook


# take number random files since there are more CR than Harlem
def data_adq(group, number):

    path = os.listdir('./data/' + group + '/')
    files = list()
    for i in range(number):
        file = random.choice(path)
        while files.count(file):
            file = random.choice(path)

        files.append(file)
    return group, files


# combine feature vectors and add group classification
def data_prep(group, files, mode):
    # change work directory for simplicity
    os.chdir('./data/' + group + '/')
    feat_vect = list()
    # loop through files getting feature vector and adding the corresponding group
    for file in files:
        mat = sio.loadmat(file)
        theta = mat.get('theta')
        class_group = np.concatenate(([int(group[11])], theta.flatten()))
        feat_vect.append(class_group)
    data_save(group, mode, feat_vect)
    # return to home directory
    os.chdir('../../..')


# save the data in a file so the same dataset could be used for further tests
def data_save(group, mode, feat_vect):
    save_to = '../' + str(group[11]) + str(mode) + '.csv'
    np.savetxt(save_to, feat_vect, delimiter=",")


# manipulate the data to remove the labels
def features_manipulation(file1, file2, file3, file4):
    features1 = pd.read_csv(file1, header=None, delimiter=",")
    features2 = pd.read_csv(file2, header=None, delimiter=",")
    features = pd.concat([features1, features2])
    features.rename(columns={0: 'class'}, inplace=True)

    # Labels are the values we want to predict
    labels = np.array(features['class'])

    # Remove the labels from the features
    features = features.drop('class', axis=1)

    # Saving feature names for later use
    feature_list = list(features.columns)
    # Convert to numpy array
    features = np.array(features)

    features3 = pd.read_csv(file3, header=None, delimiter=",")
    features4 = pd.read_csv(file4, header=None, delimiter=",")
    features_test = pd.concat([features3, features4])
    features_test.rename(columns={0: 'class'}, inplace=True)

    # Labels are the values we want to predict
    labels_test = np.array(features_test['class'])

    # Remove the labels from the features
    features_test = features_test.drop('class', axis=1)

    # Convert to numpy array
    features_test = np.array(features_test)

    return features, labels, feature_list, features_test, labels_test


# split the data into train and test
# since there is previous train and test predefined it is forced to simplify
def split_data(features_train, features_test, labels_train, labels_test):
    # Split the data into training and testing sets
    train_features, _, train_labels, _ = train_test_split(features_train, labels_train, test_size=0.0001,
                                                          random_state=32)
    _, test_features, _, test_labels = train_test_split(features_test, labels_test, test_size=0.98,
                                                        random_state=32)
    # print('Training Features Shape:', train_features.shape)
    # print('Training Labels Shape:', train_labels.shape)
    # print('Testing Features Shape:', test_features.shape)
    # print('Testing Labels Shape:', test_labels.shape)
    return train_features, test_features, train_labels, test_labels


def support_vec(train_features, test_features, train_labels, test_labels, gamma):
    # Create a svm Classifier
    clf = svm.SVC(kernel='rbf', gamma=360, degree=7)

    # Train the model using the training sets
    clf.fit(train_features, train_labels)

    # Predict the response for test dataset
    predictions = clf.predict(test_features)

    accuracy = accuracy_score(test_labels, predictions)
    # print('Accuracy:', round(accuracy, 8))

    return round(accuracy, 6)


def array_to_Excel_col(array, row, col, ws):
    for i in range(0, len(array)):
        e = ws.cell(row=row + i, column=col)
        e.value = array[i]


def conf_to_Excel(confs, row, col, ws):
    i = 0
    for conf in confs:
        array_to_Excel_col(conf, row + i, col, ws)
        i += 10


if __name__ == '__main__':
    n_test = [41,51,40,44,44]
    n_train = [126,121,125,121,114]
    #gamma= ['scale','auto',0.1,1,10,50,100,500]
    gamma = ['scale', 'auto']

    partitions = ['partition1/','partition2/','partition3/','partition4/', 'partition5/']

    wb = load_workbook('acc_results.xlsx')
    ws = wb['SVM final']
    for k in range(0, len(partitions)):
        print(k)
        j = 0
        accs = []
        for i in range (0,len(gamma)):
            accs.append([])

        for acc in accs:
            for i in range(0, 6):

                group_class_train, mat_files = data_adq(partitions[k]+'1-cr_train', n_train[k])
                data_prep(group_class_train, mat_files, 'train')
                group_class_test, mat_files = data_adq(partitions[k]+'1-cr_test', n_test[k])
                data_prep(group_class_test, mat_files, 'test')

                group_class2, mat_files2 = data_adq(partitions[k]+'2-hr_train', n_train[k])
                data_prep(group_class2, mat_files2, 'train')
                group_class2_test, mat_files2 = data_adq(partitions[k]+'2-hr_test', n_test[k])
                data_prep(group_class2_test, mat_files2, 'test')

                feature_noclass, label_class, feature_list, features_test, labels_test = features_manipulation(
                    './data/' + partitions[k] + '1train.csv', './data/' + partitions[k] + '2train.csv', './data/' + partitions[k] +'1test.csv', './data/' + partitions[k] +'2test.csv')
                train_features, test_features, train_labels, test_labels = split_data(feature_noclass, features_test,
                                                                                      label_class, labels_test)
                accuracy = support_vec(train_features, test_features, train_labels, test_labels, gamma[j])
                acc.append(accuracy)
            j += 1

        print(accs)
        row = 3
        col = 2 + k
        conf_to_Excel(accs, row, col, ws)

    wb.save('acc_results.xlsx')

